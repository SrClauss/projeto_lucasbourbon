import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, WebDriverException
import datetime
import sys
import os
import queue
import threading
import time
import traceback
import json

# Configurações globais
EXPIRATION_DATE = datetime.datetime(2025, 7, 17, 20, 0, 0, tzinfo=datetime.timezone.utc)
HEADERS = [
    "code", "name", "pricing", "discount", "pricing_with", "confins_tax", 
    "confins_value", "difalst_tax", "difalst_value", "fecop_tax", "fecop_value",
    "icmi_value", "icms_tax", "icms_value", "ipi_tax", "ipi_value", "pis_tax",
    "pis_value", "st_tax", "st_value", "weight", "status", "country_of_origin",
    "customs_tariff", "possibility_to_return"
]
header_labels ={
    "code": "Código",    
    "name": "Nome",
    "pricing": "Preço",
    "discount": "Desconto",
    "pricing_with": "Preço com Impostos",
    "cofins_tax": "Cofins",
    "cofins_value": "Cofins Valor",
    "difalst_tax": "Difal ST",
    "difalst_value": "Difal ST Valor",
    "fecop_tax": "Fecop",
    "fecop_value": "Fecop Valor",
    "icmi_value": "ICMI Valor",
    "icms_tax": "ICMS",
    "icms_value": "ICMS Valor",
    "ipi_tax": "IPI",
    "ipi_value": "IPI Valor",
    "pis_tax": "PIS",
    "pis_value": "PIS Valor",
    "st_tax": "ST",
    "st_value": "ST Valor",
    "weight": "Peso",
    "status": "Status",
    "country_of_origin": "País de Origem",
    "customs_tariff": "Tarifa Aduaneira",
    "possibility_to_return": "Possibilidade de Devolução"
}
MAX_WORKER_RETRIES = 3  # Máximo de tentativas para reiniciar um worker
WORKER_RESTART_DELAY = 5  # Segundos entre tentativas de reiniciar worker
with open('config.json', 'r') as f:
    config = json.load(f)

NUM_WORKERS = config["num_workers"]


def login():
    """Configura e inicia uma instância do Chrome com logs desativados."""
    options = ChromeOptions()
    options.add_argument('--log-level=3')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1200,800")  # Tamanho fixo para estabilidade

    driver = webdriver.Chrome(options=options)
    driver.get("https://ctshoponline.atlascopco.com/pt-BR/login")
    
    try:
        print("Aceitando cookies...")
        onetrust_accept_btn_handler = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler")))
        onetrust_accept_btn_handler.click()
        
        print("Clicando em 'Conecte-se'...")
        conecte_se_btn = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Conecte-se')]")))
        conecte_se_btn.click()
        
        print("Inserindo email...")
        input_email = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@type='email']")))
        input_email.send_keys("vendas@borbon.com.br")
        
        print("Submetendo email...")
        input_submit = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@type='submit']")))
        input_submit.click()
        
        print("Inserindo senha...")
        input_senha = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@type='password']")))
        input_senha.send_keys("Brb2025!")
        
        print("Clicando em 'Entrar'...")
        button_entrar = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, "idSIButton9")))
        button_entrar.click()
        
        print("Recusando permanecer conectado...")
        button_no = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, "idBtn_Back")))
        button_no.click()
        
        # Verificação se o login foi bem-sucedido
        print("Verificando login...")
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, "//p[contains(., 'Welcome') and .//b[text()='Vendas']]"))
        )
        print("Login bem-sucedido!")
        return driver
    except Exception as e:
        print(f"Erro durante o login: {e}")
        print(traceback.format_exc())
        driver.quit()
        return None

def search_product(driver, term):
    """Busca informações de um produto usando um driver já logado."""
    try:
        print(f"Buscando produto: {term}")
        driver.get(f"https://ctshoponline.atlascopco.com/en-GB/products/{term}")
        
        # Localizadores
        locators = {
            "product_name": (By.XPATH, "//h1[@class='mt-2']"),
            "resource_not_found": (By.XPATH, "//h2[contains(., 'The server cannot find the requested resource.')]"),
            "no_longer_available": (By.XPATH, "//*[contains(text(), 'The product is no longer available')]"),
            "cannot_add": (By.XPATH, "//h5[contains(., 'Product cannot be added to cart')]")
        }
        
        # Espera pelo nome do produto ou página não encontrada
        try:
            WebDriverWait(driver, 10).until(
                EC.any_of(
                    EC.presence_of_element_located(locators["product_name"]),
                    EC.presence_of_element_located(locators["resource_not_found"])
                )
            )
        except TimeoutException:
            print(f"  -> Tempo esgotado para carregar a página de {term}")
            return {"code": term, "status": "Tempo Esgotado"}
        
        if driver.find_elements(*locators["resource_not_found"]):
            print(f"  -> Recurso não encontrado (404): {term}")
            return {"code": term, "status": "Não Encontrado"}
        
        # Produto encontrado, continua a extração de dados
        product = {"code": term, "name": driver.find_element(*locators["product_name"]).text, "status": "Disponível"}
        print(f"  -> Produto encontrado: {product['name']}")

        try:
            # Extrai preços
            print("  Extraindo informações de preço...")
            button_pricing = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Pricing')]")))
            driver.execute_script("arguments[0].click();", button_pricing)

            WebDriverWait(driver, 15).until(
                EC.text_to_be_present_in_element((By.XPATH, "(//div[@role='tabpanel']//td)[1]"), "BRL")
            )
            
            tds = [td.text for td in driver.find_elements(By.XPATH, "//div[@role='tabpanel']//td")]
            product["pricing"] = tds[0].split(" ")[1]
            product["discount"] = "0" if tds[1] == "-" else tds[1]
            product["pricing_with"] = tds[2].split(" ")[1]
            
            # Extrai impostos
            print("  Extraindo informações de impostos...")
            button_taxes = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Taxes')]")))
            driver.execute_script("arguments[0].click();", button_taxes)
            
            table = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "table")))
            
            cells = [cell.text for cell in table.find_elements(By.XPATH, ".//td[@data-cy='informationTableCell']")]
            
        except (TimeoutException, StaleElementReferenceException):
            # Se falhar ao extrair dados, verifica se está indisponível
            if driver.find_elements(*locators["no_longer_available"]) or driver.find_elements(*locators["cannot_add"]):
                print(f"  -> Produto {term} está indisponível (detectado durante extração)")
                product["status"] = "Indisponível"
                return product
            else:
                raise  # Relança o erro se não for devido à indisponibilidade

        
        # Processamento robusto de dados fiscais
        tax_data = {
            "confins": cells[1] if len(cells) > 1 else "",
            "difalst": cells[3] if len(cells) > 3 else "",
            "fecop": cells[5] if len(cells) > 5 else "",
            "icmi": cells[7] if len(cells) > 7 else "",
            "icms": cells[9] if len(cells) > 9 else "",
            "ipi": cells[11] if len(cells) > 11 else "",
            "pis": cells[13] if len(cells) > 13 else "",
            "st": cells[15] if len(cells) > 15 else "",
        }
        
        # Função auxiliar para processar valores fiscais
        def parse_tax_value(value_str):
            if "% (BRL " in value_str:
                parts = value_str.split("% (BRL ")
                return parts[0], parts[1].replace(")", "")
            elif "BRL " in value_str:
                return "", value_str.split("BRL ")[1]
            else:
                return "", value_str

        # Atribuição dos valores
        product["confins_tax"], product["confins_value"] = parse_tax_value(tax_data["confins"])
        product["difalst_tax"], product["difalst_value"] = parse_tax_value(tax_data["difalst"])
        product["fecop_tax"], product["fecop_value"] = parse_tax_value(tax_data["fecop"])
        _, product["icmi_value"] = parse_tax_value(tax_data["icmi"])
        product["icms_tax"], product["icms_value"] = parse_tax_value(tax_data["icms"])
        product["ipi_tax"], product["ipi_value"] = parse_tax_value(tax_data["ipi"])
        product["pis_tax"], product["pis_value"] = parse_tax_value(tax_data["pis"])
        product["st_tax"], product["st_value"] = parse_tax_value(tax_data["st"])
        
        try:
            # Informações do produto
            print("  Extraindo informações do produto...")
            button_info = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Product information')]")))
            driver.execute_script("arguments[0].click();", button_info)

            # Espera a tabela de informações do produto carregar
            table = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//div[@role='tabpanel']//table")))

            # Itera sobre as linhas da tabela para extrair os dados de forma robusta
            trs = table.find_elements(By.TAG_NAME, "tr")
            for tr in trs:
                tds = tr.find_elements(By.TAG_NAME, "td")
                try:
                    label = tds[0].text
                    value = tds[1].text
                    
                    if label == "Country of origin":
                        product["country_of_origin"] = value
                    elif label == "Customs Tariff":
                        product["customs_tariff"] = value
                    elif label == "Weight":
                        product["weight"] = value
                    elif label == "Possibility to return":
                        product["possibility_to_return"] = value
                except IndexError:
                    # Linha vazia ou malformada, ignora
                    pass
        except (TimeoutException, StaleElementReferenceException):
            print("  -> Falha ao extrair informações adicionais do produto (pode não ser crítico).")
            pass  # Ignora falhas na extração de informações adicionais

        # Verificação final de disponibilidade após a coleta de dados
        if driver.find_elements(*locators["no_longer_available"]) or driver.find_elements(*locators["cannot_add"]):
            print(f"  -> Produto {term} está indisponível (detectado após extração)")
            product["status"] = "Indisponível"
        
        print(f"  -> Dados coletados para {term}")
        return product
        
    except (TimeoutException, StaleElementReferenceException) as e:
        print(f"  -> Erro temporário durante busca de {term}: {type(e).__name__}")
        raise  # Relança para tratamento específico
        
    except Exception as e:
        print(f"  -> ERRO durante processamento de {term}:")
        print(traceback.format_exc())
        return None

def worker(worker_id, driver, code_queue, results_queue, errors_queue, stop_event):
    """Processa códigos da fila usando um driver persistente."""
    print(f"Worker {worker_id}: Iniciado")
    
    while not stop_event.is_set():
        try:
            # Tenta obter um item da fila com timeout
            code, row_number = code_queue.get(timeout=10)
            print(f"Worker {worker_id}: Processando código {code} (linha {row_number})")
            
            try:
                # Tenta buscar o produto
                product_data = search_product(driver, code)
                
                if product_data:
                    # Adiciona o código aos dados para garantir que ele seja salvo
                    product_data["code"] = code
                    results_queue.put({'row_number': row_number, 'data': product_data})
                    print(f"Worker {worker_id}: Sucesso com código {code}")
                else:
                    # Produto não encontrado, envia um resultado vazio para marcar a linha
                    print(f"Worker {worker_id}: Produto {code} não encontrado ou indisponível")
                    results_queue.put({'row_number': row_number, 'data': {'code': code, 'name': 'NÃO ENCONTRADO'}})

            except (TimeoutException, StaleElementReferenceException):
                # Erros temporários - recoloca o código na fila
                print(f"Worker {worker_id}: Erro temporário com {code} - Recolocando na fila de erros")
                errors_queue.put((code, row_number))
                
                # Reinicia o driver para limpar estado
                print(f"Worker {worker_id}: Reiniciando driver...")
                try:
                    driver.quit()
                except Exception as e:
                    print(f"Worker {worker_id}: Exceção ao fechar driver: {e}")
                    pass
                
                driver = login()
                if not driver:
                    print(f"Worker {worker_id}: Falha crítica no relogin. Saindo...")
                    break

            except Exception as e:
                print(f"Worker {worker_id}: Erro inesperado com {code}:")
                print(traceback.format_exc())
            
            finally:
                # Garante que a tarefa seja marcada como concluída
                code_queue.task_done()
    
        except queue.Empty:        
            # Fila vazia por mais de 10 segundos - finaliza worker
            print(f"Worker {worker_id}: Sem códigos por 10 segundos. Finalizando...")
            break
            
    # Limpeza final
    try:
        driver.quit()
    except Exception as e:
        print(f"Worker {worker_id}: Exceção na limpeza final do driver: {e}")
        pass
    print(f"Worker {worker_id}: Finalizado")

def main():
    """Função principal com gerenciamento de filas e workers."""
    # Verificação de data de expiração
    if datetime.datetime.now(datetime.timezone.utc) > EXPIRATION_DATE:
        print("ERRO: O script expirou em 15/07/2025 e não pode mais ser executado.")
        sys.exit(1)
    
    # --- Lógica de Leitura e Preparação Centralizada ---
    input_filepath = 'lista.xlsx'
    print(f"Carregando e preparando o arquivo '{input_filepath}'...")
    try:
        workbook = openpyxl.load_workbook(input_filepath)
        sheet = workbook.active
    except FileNotFoundError:
        print(f"ERRO: Arquivo '{input_filepath}' não encontrado.")
        sys.exit(1)

    # Mapeia cabeçalhos para índices de coluna para escrita eficiente
    current_headers = [cell.value for cell in sheet[1]]
    header_to_col = {header: i + 1 for i, header in enumerate(current_headers)}

    # Adiciona novos cabeçalhos se não existirem
    new_headers_added = False
    for header_key in HEADERS:
        label = header_labels.get(header_key, header_key)
        if label not in header_to_col:
            new_headers_added = True
            new_col_idx = len(header_to_col) + 1
            sheet.cell(row=1, column=new_col_idx, value=label)
            header_to_col[label] = new_col_idx                
    
    if new_headers_added:
        print("Novos cabeçalhos adicionados à planilha.")
        workbook.save(input_filepath)

    # Identifica códigos a processar
    codes_to_process = []
    total_codes = 0
    name_col_idx = header_to_col.get(header_labels["name"]) # Coluna 'Nome' para verificar se já foi processado

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=1), start=2):
        code_cell = row[0]
        if code_cell.value:
            total_codes += 1
            code = str(code_cell.value).zfill(10)
            # Processa se a célula do nome estiver vazia
            if not sheet.cell(row=row_idx, column=name_col_idx).value:
                codes_to_process.append((code, row_idx))

    processed_count = total_codes - len(codes_to_process)
    print(f"{processed_count} códigos já processados. {len(codes_to_process)} a processar.")
    
    if not codes_to_process:
        print("Nenhum código novo para processar.")
        return

    # Cria filas
    code_queue = queue.Queue()
    results_queue = queue.Queue()
    errors_queue = queue.Queue()
    stop_event = threading.Event()
    
    # Preenche fila de códigos
    for item in codes_to_process:
        code_queue.put(item)

    # Cria e inicia workers com tratamento robusto
    active_workers = []
    worker_threads = {}
    
    # Função para iniciar um worker
    def start_worker(worker_id):
        print(f"\nIniciando worker {worker_id}...")
        driver = login()
        if driver:
            t = threading.Thread(
                target=worker,
                args=(worker_id, driver, code_queue, results_queue, errors_queue, stop_event),
                daemon=True
            )
            t.start()
            worker_threads[worker_id] = t
            active_workers.append(worker_id)
            print(f"Worker {worker_id} iniciado com sucesso")
            return True
        return False
    
    # Tenta iniciar todos os workers com múltiplas tentativas
    for i in range(NUM_WORKERS):
        worker_id = i + 1
        attempts = 0
        success = False
        
        while not success and attempts < MAX_WORKER_RETRIES:
            success = start_worker(worker_id)
            if not success:
                attempts += 1
                print(f"Falha ao iniciar worker {worker_id}, tentativa {attempts}/{MAX_WORKER_RETRIES}")
                time.sleep(WORKER_RESTART_DELAY)
        
        if not success:
            print(f"CRÍTICO: Não foi possível iniciar worker {worker_id} após {MAX_WORKER_RETRIES} tentativas")

    # Processa resultados e erros
    save_counter = 0
    SAVE_INTERVAL = NUM_WORKERS * 5  # Salva a cada 5 itens por worker
    try:
        print(f"Intervalo de salvamento dinâmico configurado para {SAVE_INTERVAL} itens.")
        while not code_queue.empty() or any(t.is_alive() for t in worker_threads.values()):
            # Verifica e reinicia workers que falharam
            for worker_id in list(active_workers): # Itera sobre uma cópia
                if not worker_threads[worker_id].is_alive():
                    print(f"Worker {worker_id} morreu. Tentando reiniciar...")
                    active_workers.remove(worker_id)
                    attempts = 0
                    success = False
                    
                    while not success and attempts < MAX_WORKER_RETRIES:
                        success = start_worker(worker_id)
                        if not success:
                            attempts += 1
                            print(f"Falha ao reiniciar worker {worker_id}, tentativa {attempts}/{MAX_WORKER_RETRIES}")
                            time.sleep(WORKER_RESTART_DELAY)
                    
                    if not success:
                        print(f"CRÍTICO: Não foi possível reiniciar worker {worker_id}")
            
            # Processa resultados bem-sucedidos
            while not results_queue.empty():
                result = results_queue.get()
                row_num = result["row_number"]
                data = result["data"]
                
                print(f"Salvando dados da linha {row_num}...")
                # Itera sobre os cabeçalhos definidos para garantir a ordem e a presença
                for header_key in HEADERS:
                    label = header_labels.get(header_key)
                    if label in header_to_col:
                        col_idx = header_to_col[label]
                        value = data.get(header_key, "")  # Usa .get para evitar erros se a chave não existir
                        sheet.cell(row=row_num, column=col_idx, value=value)
                
                save_counter += 1
                if save_counter >= SAVE_INTERVAL:
                    print(f"Salvando progresso no arquivo '{input_filepath}'...")
                    workbook.save(input_filepath)
                    save_counter = 0

                results_queue.task_done()
            
            # Processa erros (recoloca códigos na fila)
            while not errors_queue.empty():
                item = errors_queue.get()
                print(f"Recolocando código {item[0]} (linha {item[1]}) na fila")
                code_queue.put(item)
                errors_queue.task_done()
            
            # Verifica se o processamento terminou
            if stop_event.is_set() and code_queue.empty():
                break
                
            time.sleep(0.2)  # Pausa curta para evitar uso excessivo de CPU e manter responsividade
            
    except KeyboardInterrupt:
        print("\nInterrupção recebida. Finalizando...")
        stop_event.set()
    
    # Aguarda conclusão da fila de códigos
    code_queue.join()
    print("Todos os códigos processados.")
    
    # Garante processamento final
    print("Salvamento final do arquivo...")
    workbook.save(input_filepath)
    print("Processo concluído com sucesso.")

if __name__ == "__main__":
    main()
    
    
